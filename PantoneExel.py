# -*- coding: utf-8 -*-
import sqlite3
import tkinter as tk
from tkinter import *
from tkinter import filedialog
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

# создание соединения с базой данных
conn = sqlite3.connect('pantone_colors.db')
cursor = conn.cursor()

win = tk.Tk()
win.title("Pantone калькулятор")
#win.geometry("527x280+200+200")
win.iconbitmap('D:/Python/TIFF_RASTR/PantoneKalkulator/pantone.ico')
win.geometry("430x280+200+200")
win.resizable(False, False)
win.config(bg='#c8edd0')
Entry1 = tk.Entry(width=5, relief=tk.GROOVE, justify='center', font=('Arial', 11, 'bold'))
Entry1.grid(row=0, column=1, columnspan=2, stick='ns,ew')
Entry2 = tk.Entry(width=12, relief=tk.GROOVE, justify='left', font=('Arial', 11, 'bold'))
Entry2.grid(row=0, column=3, columnspan=2, stick='nse')

# создание полей Entry для отображения информации об чернилах
var = IntVar()
var.set(1)
rounding_precision = 1
rb1 = tk.Radiobutton(text='0.1', variable=var, value=1, width=3)
rb1 .grid(row=8, column=2, stick='ns', padx=(2, 2), pady=(2, 2))
rb2 = tk.Radiobutton(text='0.02', variable=var, value=2, width=3)
rb2 .grid(row=9, column=2, padx=(2, 2), pady=(2, 2))
rb3 = tk.Radiobutton(text='0.003', variable=var, value=3, width=3)
rb3.grid(row=10, column=2, padx=(2, 2), pady=(2, 2))
mass1 = tk.Entry(width=18, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2, justify='center') 
mass1.grid(row=2, column=3, columnspan=2, stick='ew')
mass2 = tk.Entry(width=18, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2, justify='center') 
mass2.grid(row=3, column=3, columnspan=2, stick='ew')
mass3 = tk.Entry(width=18, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2, justify='center')
mass3.grid(row=4, column=3, columnspan=2, stick='ew')
mass4 = tk.Entry(width=18, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2, justify='center')
mass4.grid(row=5, column=3, columnspan=2, stick='ew')


label_1 = tk.Label(win, text='Pantone:', font=('Arial', 10, 'bold'), padx=1, pady=5, bg='#B3B3B0', relief=tk.RAISED, bd=4)
label_1.grid(row=0, column=0, stick='ew')

label_2 = tk.Label(win, text='Масса краски(..)', font=('Arial', 10, 'bold'), padx=1, pady=5, bg='#B3B3B0', relief=tk.RAISED, bd=4)
label_2.grid(row=0, column=3)

label_3 = tk.Label(win, text='Базовые цвета ', font=('Arial', 10, 'bold'), padx=5, pady=5, fg='blue', relief=tk.RAISED, bd=3)
label_3.grid(row=1, column=0, columnspan=2, stick='ew')

label_4 = tk.Label(win, text='%', font=('Arial', 10, 'bold'), padx=5, pady=5, fg='blue', relief=tk.RAISED, bd=3)
label_4.grid(row=1, column=2, columnspan=1, stick='ew')

label_5 = tk.Label(win, text='Масса основных цветов(..)', font=('Arial', 10, 'bold'), padx=5, pady=5, fg='blue', relief=tk.RAISED, bd=3)
label_5.grid(row=1, column=3, columnspan=2, stick='ew')


ink1 = tk.Entry(width=18, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2, justify='center') 
ink1.grid(row=2, column=0, columnspan=2, stick='ew')
ink2 = tk.Entry(width=18, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2, justify='center') 
ink2.grid(row=3, column=0, columnspan=2, stick='ew')
ink3 = tk.Entry(width=18, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2, justify='center')
ink3.grid(row=4, column=0, columnspan=2, stick='ew')
ink4 = tk.Entry(width=18, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2, justify='center')
ink4.grid(row=5,column=0, columnspan=2, stick='ew')
percent1 = tk.Entry(width=4, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2) 
percent1.grid(row=2, column=2, stick='ew')
percent2 = tk.Entry(width=4, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2) 
percent2.grid(row=3, column=2, stick='ew')
percent3 = tk.Entry(width=4, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2) 
percent3.grid(row=4, column=2, stick='ew')
percent4 = tk.Entry(width=4, font=('Arial', 13, 'bold'), relief=tk.RAISED, bd=2)
percent4.grid(row=5, column=2, stick='ew')

simple_color = tk.Label(win, width=20)
simple_color.grid(row=9, column=0, columnspan=2, rowspan=2, stick='wnse', padx=(5, 5), pady=(5, 5))
#def browse_file():
     

color = tk.Button(width=8, text='Save', font=('Arial', 10, 'bold'), relief=tk.GROOVE, bd=5)
color.grid(row=8, column=0)
# Словарь с соответствиями текстовых строк и значений цветов
color_mapping = {
    'Yellow': '#FDE100',
    'Black': '#2F2C27',
    'Warm Red': '#DE5A4A',
    'Green': '#00A079',
    'Orange 021': '#E5801C',
    'Process Blue': '#008BCC',
    'Red 032': '#DD5354',
    'Reflex Blue': '#263F8C',
    'Rubine Red': '#D12368',
    'Purple': '#A3428F',
    'Rhodamine Red': '#CB4891',
    'Violet': '#4A3687',
    'Blue 072': '#2F408E',
    'Yellow 012': '#F7D917',
    'Trans.White': '#B3B3B0'
}

# Очистка всех полей
def clear_all_entries():
    percent1.delete(0, tk.END)
    percent2.delete(0, tk.END)
    percent3.delete(0, tk.END)
    percent4.delete(0, tk.END)
    Entry2.delete(0, tk.END)
    Entry1.delete(0, tk.END)
    ink1.delete(0, tk.END)
    ink2.delete(0, tk.END)
    ink3.delete(0, tk.END)
    ink4.delete(0, tk.END)
    mass1.delete(0, tk.END)
    mass2.delete(0, tk.END)
    mass3.delete(0, tk.END)
    mass4.delete(0, tk.END)
    simple_color.configure(bg="white")

clean = tk.Button(width=8, text='Clean', font=('Arial', 10, 'bold'), relief=tk.GROOVE, bd=5, command=clear_all_entries)
clean.grid(row=8, column=1, padx=(5, 5), pady=(5, 5))



# функция вычисления 
def calculate_ink():
    # Получаем значения массы краски и процентного содержания каждого основного цвета
    ink_mass = float(Entry2.get())
    percent1_value = float(percent1.get()) / 100
    percent2_value = float(percent2.get()) / 100
    
    try:
        percent3_value = float(percent3.get()) / 100
    except ValueError:
        percent3_value = 0.0
    
    try:
        percent4_value = float(percent4.get()) / 100
    except ValueError:
        percent4_value = 0.0


    # Вычисляем массы каждого основного цвета
    ink1_mass = ink_mass * percent1_value
    ink2_mass = ink_mass * percent2_value
    ink3_mass = ink_mass * percent3_value
    ink4_mass = ink_mass * percent4_value

    # Вычисляем общую массу краски
    total_ink_mass = ink1_mass + ink2_mass + ink3_mass + ink4_mass

    if var.get() ==1:
      rounding_precision = 1
    elif var.get() ==2:
      rounding_precision = 2
    elif var.get() ==3:
      rounding_precision = 3
    # Выводим результаты в соответствующие виджеты Entry
    mass1.delete(0, END)
    if ink1_mass != 0:
        mass1.insert(0, round(ink1_mass, rounding_precision))
    mass2.delete(0, END)
    if ink2_mass != 0:
        mass2.insert(0, round(ink2_mass, rounding_precision))
    mass3.delete(0, END)
    if ink3_mass != 0:
        mass3.insert(0, round(ink3_mass, rounding_precision))
    mass4.delete(0, END)
    if ink4_mass != 0:
        mass4.insert(0, round(ink4_mass, rounding_precision)) 

# Создаем кнопку пересчета
Entry2.bind('<Return>', lambda event: calculate_ink())


calculate = tk.Button(width=8, text='Calculate', font=('Arial', 10, 'bold'), relief=tk.GROOVE, bd=5, command=calculate_ink)
calculate.grid(row=8, column=3, padx=(5, 5), pady=(5, 5))
def update_entry_values(event):
    global mass1, mass2, mass3, mass4, Entry2
    global p1, p2, p3, p4
    
    p1 = float(percent1.get()) if percent1.get() else 0
    p2 = float(percent2.get()) if percent2.get() else 0
    p3 = float(percent3.get()) if percent3.get() else 0
    p4 = float(percent4.get()) if percent4.get() else 0
    
    # Получаем значения из всех полей ввода
    m1 = float(mass1.get()) if mass1.get() else None
    m2 = float(mass2.get()) if mass2.get() else None
    m3 = float(mass3.get()) if mass3.get() else None
    m4 = float(mass4.get()) if mass4.get() else None
    
    # Рассчитываем новое значение в поле Entry2
    total_mass = sum([m for m in [m1, m2, m3, m4] if m is not None])
    if total_mass != 0:
        if event.widget == mass1:
            percent = p1
        elif event.widget == mass2:
            percent = p2
        elif event.widget == mass3:
            percent = p3
        elif event.widget == mass4:
            percent = p4
        new_mass = float(event.widget.get())
        Entry2_value = (new_mass * 100) / percent
        Entry2.delete(0, tk.END)
        Entry2.insert(0, str(Entry2_value))
    
    # Обновляем значения в полях mass1, mass2, mass3, mass4
    mass1.delete(0, tk.END)
    if m1 is not None:
        mass1.insert(0, str(m1))
    
    mass2.delete(0, tk.END)
    if m2 is not None:
        mass2.insert(0, str(m2))
    
    mass3.delete(0, tk.END)
    if m3 is not None:
        mass3.insert(0, str(m3))
    
    mass4.delete(0, tk.END)
    if m4 is not None:
        mass4.insert(0, str(m4))
# Подписываемся на события изменения значений полей ввода
mass1.bind('<Return>', update_entry_values)
mass2.bind('<Return>', update_entry_values)
mass3.bind('<Return>', update_entry_values)
mass4.bind('<Return>', update_entry_values)

def show_suggestions(event):
    query = Entry1.get()
    if len(query) >= 2:
        cursor.execute("SELECT name FROM new_pantone_colors1 WHERE name LIKE ? UNION "
                       "SELECT name FROM new_pantone_colors2 WHERE name LIKE ? UNION "
                       "SELECT name FROM new_pantone_colors3 WHERE name LIKE ?",
                       (f'%{query}%', f'%{query}%', f'%{query}%'))
        suggestions = cursor.fetchall()

        if suggestions:
            popup = tk.Toplevel(win)  # Создание нового окна (popup)
            popup.overrideredirect(True)
            #popup.title("Подсказка")
            #popup.iconbitmap('D:/Python/TIFF_RASTR/PantoneKalkulator/pantone.ico')
            popup.config(bg='#00A17E')
            popup.geometry("200x150+300+300")
            popup.resizable(False, False)

            listbox_popup = Listbox(popup, width=198, height=148, relief=tk.GROOVE)
            listbox_popup.pack(padx=10, pady=10)

            for suggestion in suggestions:
                listbox_popup.insert(END, suggestion[0])

            def select_suggestion_popup(event):
                selected_suggestion = listbox_popup.get(listbox_popup.curselection())
                Entry1.delete(0, END)
                Entry1.insert(END, selected_suggestion)
                popup.destroy()  # Закрыть popup окно
                fill_ink_info()  # Вызывайте вашу функцию здесь

                # Удалите следующую строку, если не хотите, чтобы Entry автоматически устанавливал фокус после выбора значения
                Entry1.focus()

            listbox_popup.bind("<<ListboxSelect>>", select_suggestion_popup)
            popup.mainloop()
        else:
            listbox.delete(0, END)
            listbox.grid_forget()
    else:
        listbox.delete(0, END)
        listbox.grid_forget()


listbox = Listbox(width=17, height=8, relief=tk.GROOVE, bg='lightblue')
listbox.grid_forget()

def select_suggestion(event):
    if listbox.curselection():
        selected_suggestion = listbox.get(listbox.curselection()[0])
        Entry1.delete(0, END)
        Entry1.insert(END, selected_suggestion)
        listbox.delete(0, END)
        listbox.grid_forget()
        # Вызывайте вашу функцию здесь

        fill_ink_info()

    # Удалите следующую строку, если не хотите, чтобы Entry автоматически устанавливал фокус после выбора значения
    Entry1.focus()


listbox.bind("<<ListboxSelect>>", select_suggestion)


# функция для заполнения информации о чернилах
def fill_ink_info(event=None):
    if win.winfo_exists():
      # Проверка наличия виджета Entry1
        name = Entry1.get()
    
  

    # очистка полей ввода и метки
    for ink_entry, percent_entry in zip([ink1, ink2, ink3, ink4], [percent1, percent2, percent3, percent4]):
        ink_entry.delete(0, tk.END)
        percent_entry.delete(0, tk.END)
    simple_color.configure(bg="white")

    # создание SQL-запросов для выборки данных из всех таблиц для данного имени
    queries = [
        """
        SELECT *
        FROM new_pantone_colors1
        WHERE name = ?
        """,
        """
        SELECT *
        FROM new_pantone_colors2
        WHERE name = ?
        """,
        """
        SELECT *
        FROM new_pantone_colors3
        WHERE name = ?
        """
    ]

    # выполнение запросов и вывод результатов, если найдено совпадение
    for query in queries:
        cursor.execute(query, (name,))
        results = cursor.fetchall()
        if results:
            row = results[0]
            hex_code = row[1]
            base_color1 = row[2]
            percent_color1 = row[3]
            base_color2 = row[4]
            percent_color2 = row[5]
            base_color3 = row[6] if len(row) > 6 else ""
            percent_color3 = row[7] if len(row) > 7 else ""
            base_color4 = row[8] if len(row) > 8 else ""
            percent_color4 = row[9] if len(row) > 9 else ""

            # обновление полей ввода и метки с полученными значениями
            ink1.insert(tk.END, base_color1)
            if ink1.get() in color_mapping:
                ink1.config(fg=color_mapping[ink1.get()])
            else:
                ink1.config(fg="black")
            percent1.insert(tk.END, percent_color1)
            ink2.insert(tk.END, base_color2)
            if ink2.get() in color_mapping:
                ink2.config(fg=color_mapping[ink2.get()])
            else:
                ink2.config(fg="black")
            percent2.insert(tk.END, percent_color2)
            ink3.insert(tk.END, base_color3)
            if ink3.get() in color_mapping:
                ink3.config(fg=color_mapping[ink3.get()])
            else:
                ink3.config(fg="black")
            percent3.insert(tk.END, percent_color3)           
            ink4.insert(tk.END, base_color4)
            if ink4.get() in color_mapping:
                ink4.config(fg=color_mapping[ink4.get()])
            else:
                ink4.config(fg="black")
            percent4.insert(tk.END, percent_color4)
            simple_color.configure(bg=hex_code)
            break
   
update_color = tk.Button(win, width=8, text='Update', font=('Arial', 10, 'bold'), relief=tk.GROOVE, bd=5)
update_color.grid(row=8, column=4, stick='wnse', padx=(5, 15), pady=(5, 5))
#Entry1.bind("<KeyRelease>", fill_ink_info)
def handle_selection(event=None):
    show_suggestions(event)
    fill_ink_info()
Entry1.bind("<KeyRelease>", show_suggestions)


def save_to_excel():
    ink1_value = ink1.get()
    ink2_value = ink2.get()
    ink3_value = ink3.get()
    ink4_value = ink4.get()
    entry1_value = Entry1.get().replace("p", "", 1)  # Удаление первой буквы "p" из значения Entry1
    entry1_replace = Entry1.get().replace("p", "Pantone ", 1)

    
    if ink4_value and ink3_value and ink2_value and ink1_value:  # Если ink1, ink2, ink3 и ink4 не пустые
        workbook = openpyxl.load_workbook('Pantone_4.xlsm')
        worksheet = workbook['Pantone 4']
              # название пантона
        worksheet['E6'] = entry1_value
            # масса
        worksheet['E8'] = mass1.get()
        worksheet['E9'] = mass2.get()
        worksheet['E10'] = mass3.get()
        worksheet['E11'] = mass4.get()

            #проценты
        worksheet['D8'] = percent1.get()
        worksheet['D9'] = percent2.get()
        worksheet['D10'] = percent3.get()
        worksheet['D11'] = percent4.get()
            # базовые цвета
        worksheet['B11'] = ink4.get()
        worksheet['B10'] = ink3.get()
        worksheet['B9'] = ink2.get()
        worksheet['B8'] = ink1.get()
            #Общая масса
        worksheet['E12'] = Entry2.get()

    elif ink3_value and ink2_value and ink1_value:  # Если ink1, ink2 и ink3 не пустые
        workbook = openpyxl.load_workbook('Pantone_3.xlsm')
        worksheet = workbook['Pantone 3']
              # название пантона
        worksheet['E6'] = entry1_value
            # масса
        worksheet['E8'] = mass1.get()
        worksheet['E9'] = mass2.get()
        worksheet['E10'] = mass3.get()
 

            #проценты
        worksheet['D8'] = percent1.get()
        worksheet['D9'] = percent2.get()
        worksheet['D10'] = percent3.get()
       
            # базовые цвета
        worksheet['B10'] = ink3.get()
        worksheet['B9'] = ink2.get()
        worksheet['B8'] = ink1.get()
            #Общая масса
        worksheet['E11'] = Entry2.get()
    elif ink1_value and ink2_value:  # Если ink1 и ink2 не пустые
        workbook = openpyxl.load_workbook('Pantone_2.xlsm')
        worksheet = workbook['Pantone 2']
              # название пантона
        worksheet['E6'] = entry1_value
            # масса
        worksheet['E8'] = mass1.get()
        worksheet['E9'] = mass2.get()
        
            #проценты
        worksheet['D8'] = percent1.get()
        worksheet['D9'] = percent2.get()
            
            # базовые цвета
        worksheet['B8'] = ink1.get()
        worksheet['B9'] = ink2.get()
       
            #Общая масса
        worksheet['E10'] = Entry2.get()
    else:
        print('Заполните поля')


    if workbook:
        worksheet = workbook.active

            # Получаем название листа
        sheet_name = entry1_replace 
    

            # Создаем объект выравнивания
        align = Alignment(horizontal='center', vertical='center')
            # определение свойств для контура ячейки
        border_style = Side(style='medium', color='000000')
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
         
        # сохранение книги Excel
         # Открытие диалогового окна сохранения файла
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"{sheet_name}.xlsx")

        if file_path:
            # Сохранение книги Excel
            workbook.save(file_path)
            print("Файл сохранен.")
        else:
            print("Отменено сохранение файла.")

        #workbook.save(f'D:/Exel_JOB_AMPG/{sheet_name}.xlsx')
update_color = tk.Button(win, width=8, text='Print', font=('Arial', 10, 'bold'), relief=tk.GROOVE, bd=5, command=save_to_excel)
update_color.grid(row=8, column=4, stick='wnse', padx=(5, 15), pady=(5, 5))
win.mainloop()

# закрытие соединения с базой данных
conn.close()




